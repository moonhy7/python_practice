# 엑셀 샘플 파일을 data 폴더에 복사 붙여넣기 하세요
# sample.xlsx
# TrafficAccident2019.xlsx
# TrafficAccident2019_seoul.xlsx

# [엑셀 파일 다루기]
# 관련 모듈 : openpyxl
# https://openpyxl.readthedocs.io/en/stable/
# 터미널 모드에서  명령 확인
# pip list
# openpyxl 목록 확인
# 없으면 설치
# pip install openpyxl
# 파이참 프로그램에서는 [File]-[Settings]

# 엑셀 구성요소
# 워크북(workbook=엑셀파일) > 워크시트(worksheet) > 셀(Cell)

# 모듈 임포트
# import openpyxl
# 모듈명 없이 바로 함수 호출
from openpyxl import *
import os

# 엑셀파일(xlsx, xls) 불러오기
# 엑셀워크북객체 = load_workbook(url, data_only=True)
print(os.getcwd()) # C:\pyclass
wb = load_workbook('data/sample.xlsx', data_only=True)
print(wb)
# <openpyxl.workbook.workbook.Workbook object at 0x000002016F7F5AF0>
print(type(wb))
# <class 'openpyxl.workbook.workbook.Workbook'>

# 워크시트 객체생성
# 워크시트객체변수 = 워크북객체[워크시트명]
ws = wb['영업사원매출']
print(ws, type(ws))
# <Worksheet "영업사원매출"> <class 'openpyxl.worksheet.worksheet.Worksheet'>

# 워크시트안의 셀 주소
# 워크시트객체['컬럼명행번호'].value
# 워크시트객체.cell(x,y).value
# x, y는 행과 열의 인덱싱으로 1부터 시작
print(ws['A1'].value) # Sap Co.
print(ws.cell(1,1).value) # Sap Co.
print(ws['B2'].value) # 경기수원대리점
print(ws.cell(2,2).value) # 경기수원대리점

# for문을 이용해서 특정 셀안의 내용 출력
print('='*50)
for row in ws['A1':'G1']:
    for cell in row:
        print(cell.value, end=' ')
# Sap Co. 대리점 영업사원 전월 금월 TEAM 총 판매수량

# ws.rows 모든행
# ws.columns 모든 열
# 값 < 셀 < 컬럼 또는 행 < 워크시트 < 워크북
print('===== 모든행 =======')
for row in ws.rows:
    print(row)

print('===== 모든컬럼 =======')
for column in ws.columns:
    print(column)

print('===== 모든행의 셀안의 값 =======')
for row in ws.rows:
    for cell in row:
        print(cell.value, end=' ')
    print()

print('===== 모든컬럼의 셀안의 값 =======')
for column in ws.columns:
    for cell in column:
        print(cell.value, end=' ')
    print('')
    print('='*50)

print('= 모든 셀안의 값을 2차원 리스트로 저장(행방향) =')
data_list = []
for row in ws.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    data_list.append(temp)
print(data_list)

# 워크북 객체 닫기
wb.close()

print(data_list[0])
print('='*50)
for a, b, c, d, e, f, g in data_list[1:]:
    print(a, b, c, d, e, f, g)

# 금월 매출액이 가장 높은 영업 사원은?
# 금월 매출액의 최대값 구하기
result_list = []
for a, b, c, d, e, f, g in data_list[1:]:
    result_list.append(e)

print('금월 매출액의 최대값 = ' , max(result_list))

for a, b, c, d, e, f, g in data_list[1:]:
    if e == max(result_list):
        print(f' 금월 매출액이 가장 높은 영업 사원은 {b} { f} 팀 {c} 사원입니다.')

# 금월 영업 실적이 전월 영업실적보다 떨어진 영업 사원 목록은?
# 금월영업실적 - 전월영업실적 < 0
print('=== 금월 영업 실적이 전월 영업실적보다 떨어진 영업 사원 목록')
for a, b, c, d, e, f, g in data_list[1:]:
    if e-d < 0:
        print(b, c, e-d)

# 퀴즈 : 경기도 지역의 영업 사원 목록 데이타만 csv 파일로 저장하기
# output/sample_result.csv


kyonggi_list = []
for a, b, c, d, e, f, g in data_list[1:]:
    if b.find('경기') != -1:
    # if '경기' in b:
        kyonggi_list.append([a, b, c, d, e, f, g])



kyonggi_list.insert(0, data_list[0])
print(kyonggi_list)
print(type(kyonggi_list))

# openpyxl , csv 충돌 오류 발생
import csv
# TypeError: load_workbook() got an unexpected keyword argument 'encoding'
# f = open('output/sample_result.csv', 'w', newline='', encoding='cp949' )
# csv_data = csv.writer(f)
# csv_data.writerows(kyonggi_list)
# f.close()

# TypeError: load_workbook() got an unexpected keyword argument 'encoding'
# with open('output/sample_result.csv', 'w', encoding='cp949', newline='') as f:
#     csv_data = csv.writer(f, delimiter=',')
#     csv_data.writerows(kyonggi_list)

print(kyonggi_list)

# pip install pandas
import pandas as pd
# 데이타프레임 생성
# 데이타프레임변수 = pd.DataFrame(2차원데이타/딕셔너리리스트, columns=필드리스트)
df = pd.DataFrame(kyonggi_list[1:], columns=kyonggi_list[0])
print(df)

# 데이타프레임 => csv 쓰기
# 데이타프레임변수.to_csv('csvURL', encoding='utf-8/cp949', index=False)
# index=False : 인덱스 컬럼 저장 여부
df.to_csv('output/sample_result.csv', encoding='utf-8', index=False)



