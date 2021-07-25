from openpyxl import *
import os
import pandas as pd

# 워크북(엑셀파일) > 워크시트 > 행과 열 > 셀 > 값
# 워크북객체 생성
# 워크북객체 = Workbook()
wb = Workbook()
print(wb, type(wb))

# 워크시트 생성
# 워크시트객체 = 워크북객체.create_sheet(시트명)
ws = wb.create_sheet('생성시트')

# 워크시트객체 활성화
# 워크시트객체 = 워크북객체.active
ws = wb.active

# 특정 셀안에 값 쓰기 1 - 셀단위
# 워크시트객체['필드명행인덱스'] = 값
ws['A1'] = 'Start1'
ws['B1'] = 'Start2'

# 특정 셀안에 값 쓰기 2 - 셀단위
# 워크시트객체.cell(x, y, 값)
ws.cell(3,3, '3행3열')
ws.cell(5,5, '5행5열')

# 행단위로 셀안에 값 쓰기
# 워크시트객체.append(리스트/튜플)
ws.append([100, 200, 300, 400])
ws.append([600, 55, 34])

# 엑셀파일에 쓰기
# 워크북객체.save(엑셀파일경로)
wb.save('output/test1.xlsx')

# 엑셀파일 닫기
# 워크북객체.close()
wb.close()

# ==============
# 2차원 리스트를 엑셀파일로 저장하기
# 5행 10열
data_list = []
for i in range(1,6):
    temp = []
    for j in range(1,11):
        temp.append('python-'+str(i)+'-'+ str(j))
    data_list.append(temp)
print(data_list)
wb = Workbook()
ws = wb.active
ws.append(['cell1', 'cell2', 'cell3', 'cell4', 'cell5'])
for row in data_list:
    ws.append(row)
ws.append(['end1', 'end2', 'end3', 'end4'])

wb.save('output/test2.xls')
wb.close()

# ==============
# 데이타 수집
# 교통안전정보관리시스템 https://tmacs.kotsa.or.kr/
# [교통사고 원인분석]-[교통사고 현황]-[지자체별 교통사고]
# TrafficAccident2019.xlsx 파일을 이용해서
# 교통사고 발생건수가 가장 높은 5대 시도의 데이타만 엑셀 파일로 저장한다.

# 워크북 객체 생성
wb = load_workbook('data/TrafficAccident2019.xlsx', data_only=True)
# 워크시트 객체 생성
ws = wb['2019']
print('\n'*5)
print(ws) # <Worksheet "2019">

# 제목 필드 리스트 생성
field_list = []
for row in ws['A2':'G2']:
    for cell in row:
        field_list.append(cell.value)

print(field_list)
# ['시도', '발생건수', '대형사고', '중대사고', None, '사망자수', '치사율(%)']
field_list[3] = '여객건'
field_list[4] = '화물건'
print(field_list)
# ['시도', '발생건수', '대형사고', '여객건', '화물건', '사망자수', '치사율(%)']

# 전체 셀의 값
data_list = []
for row in ws.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    data_list.append(temp)
print(data_list)

# 필요한 데이타만 다시 재정의
print(data_list[3:-1])
data_list = data_list[3:-1]

# 판다스의 데이타프레임 모양으로 확인하기
df = pd.DataFrame(data_list, columns=field_list)
print(df)

# 교통사고 발생건수 기준으로 정렬
temp_list = []
for a, b, c, d, e, f, g in data_list:
    temp_list.append(b)
print ('Step1', temp_list)
temp_list.sort()
print ('Step2', temp_list)
# temp_list.reverse()
temp_list = temp_list[::-1]
print ('Step3', temp_list)

# 교통사고가 많은 데이타 5개만 저장하기
result_list = []
for i in range(5):
    for a, b, c, d, e, f, g in data_list:
        if b == temp_list[i]:
            result_list.append([i+1, a, b])
print(result_list)
wb.close()

# 엑셀 파일로 저장하기
wb = Workbook()
ws = wb.active
# 워크시트 이름 변경
ws.title = '교통사고 발생건수'
ws.append(['순위', '시도', '발생건수'])
for row in result_list:
    ws.append(row)
wb.save('output/traffic_accident5.xls')
wb.close()





